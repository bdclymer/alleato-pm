#!/usr/bin/env python3
"""
Import budget data from Excel file into Supabase for a project.

Usage:
    cd backend
    source .venv/bin/activate
    PYTHONPATH="src" python scripts/import_budget_from_excel.py --file path/to/budget.xlsx --project-id 67
"""

import argparse
import os
import sys
from datetime import datetime, UTC
from typing import Optional, List, Dict
import openpyxl

# Add canonical backend src path for imports. Retired worker package paths must
# not be reintroduced; background ingestion lives under src/services.
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

# Load environment variables
from dotenv import load_dotenv

# Try multiple locations for .env
env_locations = [
    os.path.join(os.path.dirname(__file__), '..', '..', '.env'),
    os.path.join(os.path.dirname(__file__), '..', '..', '.env.local'),
    os.path.join(os.path.dirname(__file__), '..', '.env'),
    os.path.join(os.path.dirname(__file__), '..', '.env.local'),
]

for env_path in env_locations:
    if os.path.exists(env_path):
        load_dotenv(env_path)
        print(f"Loaded env from: {env_path}")
        break

from services.supabase_helpers import get_supabase_client


def read_excel_budget(file_path: str) -> Dict:
    """Read budget data from Excel file."""
    print(f"\nReading budget file: {file_path}")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    print(f"Workbook loaded. Available sheets: {wb.sheetnames}")

    # Analyze the first sheet to understand structure
    sheet = wb.active
    print(f"\nAnalyzing sheet: {sheet.title}")
    print(f"Dimensions: {sheet.dimensions}")

    # Print first 10 rows to understand structure
    print("\nFirst 10 rows of data:")
    print("-" * 80)

    data = []
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if i <= 10:
            print(f"Row {i}: {row}")
        data.append(row)

    return {
        'sheets': wb.sheetnames,
        'active_sheet': sheet.title,
        'dimensions': sheet.dimensions,
        'data': data,
        'workbook': wb
    }


def parse_budget_structure(excel_data: Dict) -> Dict:
    """Parse the Excel data to identify budget structure."""
    wb = excel_data['workbook']
    sheet = wb.active

    # Find headers
    headers = None
    header_row = None

    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        # Look for a row that contains budget-related headers
        row_str = ' '.join([str(cell).lower() for cell in row if cell])
        if any(keyword in row_str for keyword in ['cost', 'budget', 'amount', 'total', 'description', 'item']):
            headers = [cell for cell in row if cell]
            header_row = i
            print(f"\nFound headers in row {i}: {headers}")
            break

    if not headers:
        print("\nWarning: Could not identify header row automatically.")
        print("Defaulting to first row as headers.")
        headers = [cell for cell in next(sheet.iter_rows(values_only=True))]
        header_row = 1

    # Extract budget items
    budget_items = []
    for i, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if not any(row):  # Skip empty rows
            continue

        item = {}
        for j, (header, value) in enumerate(zip(headers, row)):
            if header:
                item[str(header).strip()] = value

        if item:
            budget_items.append(item)

    return {
        'headers': headers,
        'header_row': header_row,
        'items': budget_items,
        'total_items': len(budget_items)
    }


def get_or_create_cost_code(client, project_id: int, code_str: str) -> Optional[str]:
    """Get or create a cost code for the project."""
    # Parse code and description from "013120 - Vice President - Labor"
    if ' - ' in code_str:
        parts = code_str.split(' - ', 1)
        code = parts[0].strip()
        description = parts[1].strip() if len(parts) > 1 else code
    else:
        code = code_str.strip()
        description = code

    # Try to find existing cost code
    try:
        result = client.table('cost_codes').select('id').eq('code', code).eq('project_id', project_id).limit(1).execute()
        if result.data and len(result.data) > 0:
            return result.data[0]['id']
    except Exception as e:
        print(f"  Note: Error checking existing cost code: {e}")

    # Create new cost code
    try:
        result = client.table('cost_codes').insert({
            'code': code,
            'description': description,
            'project_id': project_id
        }).execute()
        if result.data and len(result.data) > 0:
            print(f"  Created cost code: {code} - {description}")
            return result.data[0]['id']
    except Exception as e:
        print(f"  Error creating cost code {code}: {e}")

    return None


def import_budget_to_database(client, project_id: int, budget_data: Dict, dry_run: bool = True) -> bool:
    """Import budget items into Supabase."""

    print(f"\n{'='*80}")
    print(f"Importing Budget for Project ID: {project_id}")
    print(f"Total items to import: {budget_data['total_items']}")
    print(f"{'='*80}\n")

    if dry_run:
        print("DRY RUN MODE - No data will be inserted\n")

    # Display sample data
    print("Sample budget items (first 5):")
    for i, item in enumerate(budget_data['items'][:5], 1):
        print(f"\n{i}. {item}")

    if dry_run:
        print(f"\n{'='*80}")
        print("Field Mapping:")
        print(f"{'='*80}")
        print("Excel Column -> Database Column")
        print("Budget Code -> cost_code_id (will create/lookup cost codes)")
        print("Original Budget -> original_budget_amount")
        print("Budget Modifications -> budget_modifications")
        print("Revised Budget -> revised_budget")
        print("Projected Budget -> (calculated)")
        print("Approved Changes -> approved_cos")
        print("Commitments -> committed_cost")
        print("Direct Costs -> direct_cost")
        print("Projected Costs -> projected_cost")
        print("Forecast to Complete -> forecast_to_complete")
        print("Pending Costs -> pending_cost_changes")
        return True

    # Actually import the data
    print(f"\n{'='*80}")
    print("Importing budget items...")
    print(f"{'='*80}\n")

    imported_count = 0
    skipped_count = 0

    for i, item in enumerate(budget_data['items'], 1):
        budget_code = item.get('Budget Code')
        if not budget_code:
            print(f"  Row {i}: Skipping - no budget code")
            skipped_count += 1
            continue

        # Get or create cost code
        cost_code_id = get_or_create_cost_code(client, project_id, budget_code)
        if not cost_code_id:
            print(f"  Row {i}: Skipping - could not create cost code for '{budget_code}'")
            skipped_count += 1
            continue

        # Map Excel columns to database fields
        budget_item = {
            'project_id': project_id,
            'cost_code_id': cost_code_id,
            'original_budget_amount': item.get('Original Budget', 0) or 0,
            'budget_modifications': item.get('Budget Modifications', 0) or 0,
            'revised_budget': item.get('Revised Budget') or None,
            'approved_cos': item.get('Approved Changes', 0) or 0,
            'committed_cost': item.get('Commitments') or None,
            'direct_cost': item.get('Direct Costs') or None,
            'projected_cost': item.get('Projected Costs') or None,
            'forecast_to_complete': item.get('Forecast to Complete') or None,
            'pending_cost_changes': item.get('Pending Costs') or None,
        }

        try:
            result = client.table('budget_items').insert(budget_item).execute()
            if result.data:
                imported_count += 1
                if imported_count % 10 == 0:
                    print(f"  Imported {imported_count} items...")
        except Exception as e:
            print(f"  Row {i}: Error importing '{budget_code}': {e}")
            skipped_count += 1

    print(f"\n{'='*80}")
    print(f"Import Complete")
    print(f"{'='*80}")
    print(f"Imported: {imported_count} items")
    print(f"Skipped: {skipped_count} items")
    print(f"{'='*80}\n")

    return imported_count > 0


def main():
    parser = argparse.ArgumentParser(
        description='Import budget data from Excel into Supabase'
    )
    parser.add_argument(
        '--file',
        required=True,
        help='Path to Excel budget file'
    )
    parser.add_argument(
        '--project-id',
        type=int,
        required=True,
        help='Project ID to associate budget with'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        default=True,
        help='Preview import without saving (default: True)'
    )
    parser.add_argument(
        '--import',
        dest='do_import',
        action='store_true',
        help='Actually import the data (disables dry-run)'
    )

    args = parser.parse_args()

    print(f"\n{'='*80}")
    print(f"Budget Import Tool")
    print(f"{'='*80}\n")

    # Read Excel file
    try:
        excel_data = read_excel_budget(args.file)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return 1

    # Parse budget structure
    print(f"\n{'='*80}")
    print("Parsing budget structure...")
    print(f"{'='*80}")

    budget_data = parse_budget_structure(excel_data)

    print(f"\nParsed {budget_data['total_items']} budget items")
    print(f"Headers: {budget_data['headers']}")

    # Connect to Supabase
    supabase = get_supabase_client()

    # Verify project exists
    try:
        project = supabase.table('projects').select('id, name').eq('id', args.project_id).single().execute()
        print(f"\nTarget project: {project.data['name']} (ID: {project.data['id']})")
    except Exception as e:
        print(f"Error: Project {args.project_id} not found: {e}")
        return 1

    # Import budget
    dry_run = not args.do_import
    success = import_budget_to_database(supabase, args.project_id, budget_data, dry_run=dry_run)

    if dry_run:
        print(f"\n{'='*80}")
        print("DRY RUN COMPLETE")
        print(f"{'='*80}")
        print("\nTo actually import this data, run with --import flag")

    return 0 if success else 1


if __name__ == '__main__':
    sys.exit(main())
