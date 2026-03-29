#!/usr/bin/env python3
"""
Convert legacy budget exports into Alleato import rows and optionally import them
directly into Supabase.

Examples:
  python3 scripts/import_legacy_budget.py --list-projects
  python3 scripts/import_legacy_budget.py --input ~/Downloads/budget-16.xlsx --convert-only
  python3 scripts/import_legacy_budget.py --input ~/Downloads/budget-16.xlsx --project 67 --import
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required. Install it with `python3 -m pip install openpyxl`."
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "output" / "spreadsheet"
ENV_CANDIDATES = (
    ROOT / ".env.local",
    ROOT / ".env",
    ROOT / "frontend" / ".env.local",
)

COST_TYPE_MAP = {
    "labor": "L",
    "expense": "X",
    "subcontract": "S",
    "material": "M",
    "equipment": "E",
    "other": "O",
    "revenue": "R",
}


@dataclass
class LegacyBudgetRow:
    cost_code: str
    cost_type: str
    description: str
    original_budget: float
    source_budget_code: str


class SupabaseRestClient:
    def __init__(self, url: str, service_key: str):
        self.base_url = url.rstrip("/") + "/rest/v1"
        self.headers = {
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }

    def request(
        self,
        path: str,
        *,
        method: str = "GET",
        params: dict[str, str] | None = None,
        payload: Any | None = None,
        prefer: str | None = None,
    ) -> tuple[Any, dict[str, str]]:
        query = ""
        if params:
            query = "?" + urllib.parse.urlencode(params, safe="(),*")
        url = f"{self.base_url}/{path}{query}"
        headers = dict(self.headers)
        if prefer:
            headers["Prefer"] = prefer
        body = None
        if payload is not None:
            body = json.dumps(payload).encode("utf-8")

        request = urllib.request.Request(url, data=body, headers=headers, method=method)
        try:
            with urllib.request.urlopen(request) as response:
                raw = response.read().decode("utf-8")
                data = json.loads(raw) if raw else None
                return data, dict(response.headers.items())
        except urllib.error.HTTPError as exc:
            details = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(
                f"Supabase {method} {path} failed with {exc.code}: {details}"
            ) from exc

    def select(self, table: str, params: dict[str, str]) -> list[dict[str, Any]]:
        data, _ = self.request(table, params=params)
        return data or []

    def insert(
        self, table: str, payload: dict[str, Any] | list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        data, _ = self.request(
            table,
            method="POST",
            payload=payload,
            prefer="return=representation",
        )
        if data is None:
            return []
        return data if isinstance(data, list) else [data]

    def patch(self, table: str, filters: dict[str, str], payload: dict[str, Any]) -> Any:
        data, _ = self.request(
            table,
            method="PATCH",
            params=filters,
            payload=payload,
            prefer="return=representation",
        )
        return data


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    for path in ENV_CANDIDATES:
        if not path.exists():
            continue
        for raw_line in path.read_text().splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            env.setdefault(key, value.strip().strip('"').strip("'"))
    env.update({key: value for key, value in os.environ.items() if value})
    return env


def get_supabase_client() -> SupabaseRestClient:
    env = load_env()
    url = env.get("SUPABASE_URL") or env.get("NEXT_PUBLIC_SUPABASE_URL")
    service_key = (
        env.get("SUPABASE_SERVICE_ROLE_KEY")
        or env.get("SUPABASE_SERVICE_KEY")
        or env.get("SUPABASE_SECRET_KEY")
    )
    if not url or not service_key:
        raise SystemExit(
            "Missing Supabase credentials. Expected SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY."
        )
    return SupabaseRestClient(url, service_key)


def normalize_cost_code(raw: str) -> str:
    value = raw.strip()
    if "-" in value:
        left, right = value.split("-", 1)
        return f"{left.zfill(2)}-{right}"

    digits_only = re.sub(r"\D", "", value)
    if digits_only and len(digits_only) == 6:
        return f"{digits_only[:2]}-{digits_only[2:]}"
    if digits_only and len(digits_only) == 5:
        return f"{digits_only[:1].zfill(2)}-{digits_only[1:]}"
    return value


def infer_cost_type(description: str) -> str:
    tail = description.split(" - ")[-1].strip().lower() if description else ""
    if tail in COST_TYPE_MAP:
        return COST_TYPE_MAP[tail]
    raise ValueError(
        f'Could not infer cost type from description "{description}". '
        "Expected the last description segment to be Labor, Expense, Subcontract, Material, Equipment, Other, or Revenue."
    )


def parse_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace("$", "").replace(",", "").strip()
    return float(cleaned) if cleaned else 0.0


def parse_legacy_workbook(path: Path) -> list[LegacyBudgetRow]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook is empty.")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    expected = "Budget Code"
    if expected not in header:
        raise ValueError(
            f'Expected first sheet to contain a "{expected}" column. Found: {header}'
        )

    index = {name: idx for idx, name in enumerate(header)}
    parsed_rows: list[LegacyBudgetRow] = []

    for row in rows[1:]:
        if not row:
            continue
        budget_code_cell = row[index["Budget Code"]] if index["Budget Code"] < len(row) else None
        if budget_code_cell is None:
            continue

        source_budget_code = str(budget_code_cell).strip()
        if not source_budget_code or source_budget_code.lower() == "grand total:":
            continue

        original_budget = parse_number(
            row[index.get("Original Budget", -1)] if "Original Budget" in index else 0
        )

        parts = source_budget_code.split(" - ", 1)
        description = parts[1].strip() if len(parts) > 1 else ""
        cost_code = normalize_cost_code(parts[0])
        cost_type = infer_cost_type(description)

        parsed_rows.append(
            LegacyBudgetRow(
                cost_code=cost_code,
                cost_type=cost_type,
                description=description,
                original_budget=original_budget,
                source_budget_code=source_budget_code,
            )
        )

    if not parsed_rows:
        raise ValueError("No budget rows found after removing totals/blank rows.")

    return parsed_rows


def write_import_csv(rows: list[LegacyBudgetRow], source_path: Path) -> Path:
    DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = DEFAULT_OUTPUT_DIR / f"{source_path.stem}-import-ready.csv"
    with output_path.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "Cost Code",
                "Cost Type",
                "Description",
                "Unit Qty",
                "UOM",
                "Unit Cost",
                "Original Budget",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "Cost Code": row.cost_code,
                    "Cost Type": row.cost_type,
                    "Description": row.description,
                    "Unit Qty": "",
                    "UOM": "",
                    "Unit Cost": "",
                    "Original Budget": f"{row.original_budget:.2f}",
                }
            )
    return output_path


def list_projects(client: SupabaseRestClient, query: str | None) -> None:
    projects = client.select(
        "projects",
        {
            "select": "id,name,project_number,archived",
            "order": "id.asc",
        },
    )
    filtered = [
        project
        for project in projects
        if not project.get("archived")
        and (
            not query
            or query.lower() in str(project.get("id", "")).lower()
            or query.lower() in str(project.get("name", "")).lower()
            or query.lower() in str(project.get("project_number", "")).lower()
        )
    ]

    if not filtered:
        print("No matching active projects found.")
        return

    for project in filtered:
        print(
            f'{project["id"]:>4}  {project.get("project_number") or "-":<18}  {project.get("name") or "(unnamed)"}'
        )


def resolve_project(client: SupabaseRestClient, raw_project: str) -> dict[str, Any]:
    if raw_project.isdigit():
        projects = client.select(
            "projects",
            {
                "select": "id,name,project_number,archived",
                "id": f"eq.{raw_project}",
            },
        )
        if not projects:
            raise ValueError(f"Project {raw_project} was not found.")
        return projects[0]

    projects = client.select(
        "projects",
        {
            "select": "id,name,project_number,archived",
            "order": "id.asc",
        },
    )
    matches = [
        project
        for project in projects
        if not project.get("archived")
        and (
            raw_project.lower() in str(project.get("name", "")).lower()
            or raw_project.lower() in str(project.get("project_number", "")).lower()
        )
    ]
    if not matches:
        raise ValueError(f'No active project matched "{raw_project}".')
    if len(matches) > 1:
        lines = [
            f'{project["id"]}: {project.get("project_number") or "-"} - {project.get("name") or "(unnamed)"}'
            for project in matches
        ]
        raise ValueError(
            f'Multiple active projects matched "{raw_project}". Use a numeric project id instead:\n'
            + "\n".join(lines)
        )
    return matches[0]


def build_cost_code_candidates(cost_code: str) -> list[str]:
    normalized = normalize_cost_code(cost_code)
    compact = normalized.replace("-", "")
    return [candidate for candidate in dict.fromkeys([normalized, compact]) if candidate]


def resolve_division_id(cost_code: str, divisions: list[dict[str, Any]]) -> str:
    prefix = cost_code.split("-", 1)[0].lstrip("0") or "0"
    for division in divisions:
        division_code = str(division.get("code", "")).strip()
        if division_code == prefix or division_code.lstrip("0") == prefix:
            return str(division["id"])
    raise ValueError(f'Could not determine division for cost code "{cost_code}".')


def ensure_cost_code(
    client: SupabaseRestClient,
    row: LegacyBudgetRow,
    divisions: list[dict[str, Any]],
    dry_run: bool,
) -> tuple[str, list[str]]:
    warnings: list[str] = []
    for candidate in build_cost_code_candidates(row.cost_code):
        existing = client.select(
            "cost_codes",
            {
                "select": "id,title",
                "id": f"eq.{candidate}",
                "limit": "1",
            },
        )
        if existing:
            existing_id = str(existing[0]["id"])
            if existing_id != row.cost_code:
                warnings.append(
                    f'Using existing cost code "{existing_id}" for source "{row.source_budget_code}".'
                )
            return existing_id, warnings

    division_id = resolve_division_id(row.cost_code, divisions)
    if dry_run:
        warnings.append(f'Would create missing cost code "{row.cost_code}".')
        return row.cost_code, warnings

    client.insert(
        "cost_codes",
        {
            "id": row.cost_code,
            "title": row.description or row.cost_code,
            "division_id": division_id,
            "status": "active",
        },
    )
    warnings.append(f'Created missing cost code "{row.cost_code}".')
    return row.cost_code, warnings


def ensure_project_cost_code(
    client: SupabaseRestClient,
    project_id: int,
    cost_code_id: str,
    cost_type_id: str,
    dry_run: bool,
) -> list[str]:
    warnings: list[str] = []
    existing = client.select(
        "project_cost_codes",
        {
            "select": "id,is_active",
            "project_id": f"eq.{project_id}",
            "cost_code_id": f"eq.{cost_code_id}",
            "cost_type_id": f"eq.{cost_type_id}",
            "limit": "1",
        },
    )
    if existing:
        if existing[0].get("is_active") is False:
            if dry_run:
                warnings.append(
                    f'Would reactivate project cost code "{cost_code_id}" for project {project_id}.'
                )
                return warnings
            client.patch(
                "project_cost_codes",
                {"id": f'eq.{existing[0]["id"]}'},
                {"is_active": True},
            )
            warnings.append(
                f'Reactivated project cost code "{cost_code_id}" for project {project_id}.'
            )
        return warnings

    if dry_run:
        warnings.append(
            f'Would add project cost code "{cost_code_id}" to project {project_id}.'
        )
        return warnings

    client.insert(
        "project_cost_codes",
        {
            "project_id": project_id,
            "cost_code_id": cost_code_id,
            "cost_type_id": cost_type_id,
            "is_active": True,
        },
    )
    warnings.append(
        f'Added project cost code "{cost_code_id}" to project {project_id}.'
    )
    return warnings


def update_project_budget_summary(
    client: SupabaseRestClient,
    project_id: int,
) -> tuple[float, str | None]:
    budget_lines = client.select(
        "budget_lines",
        {
            "select": "original_amount",
            "project_id": f"eq.{project_id}",
            "limit": "10000",
        },
    )
    total = sum(parse_number(row.get("original_amount")) for row in budget_lines)
    try:
        client.patch(
            "projects",
            {"id": f"eq.{project_id}"},
            {"budget": total},
        )
        return total, None
    except RuntimeError as exc:
        return total, str(exc)


def import_rows(
    client: SupabaseRestClient,
    project_id: int,
    rows: list[LegacyBudgetRow],
    dry_run: bool,
) -> dict[str, Any]:
    divisions = client.select(
        "cost_code_divisions",
        {
            "select": "id,code,title",
            "is_active": "eq.true",
            "order": "code.asc",
        },
    )
    if not divisions:
        raise ValueError("No active cost code divisions found.")

    cost_types = {
        str(item["code"]).upper(): str(item["id"])
        for item in client.select("cost_code_types", {"select": "id,code"})
    }

    imported_count = 0
    skipped_duplicates = 0
    warnings: list[str] = []

    for index, row in enumerate(rows, start=2):
        cost_type_id = cost_types.get(row.cost_type.upper())
        if not cost_type_id:
            raise ValueError(
                f'Cost type "{row.cost_type}" is not configured in Supabase (source row "{row.source_budget_code}").'
            )

        cost_code_id, cost_code_warnings = ensure_cost_code(client, row, divisions, dry_run)
        warnings.extend(f"Row {index}: {warning}" for warning in cost_code_warnings)

        project_cost_code_warnings = ensure_project_cost_code(
            client, project_id, cost_code_id, cost_type_id, dry_run
        )
        warnings.extend(f"Row {index}: {warning}" for warning in project_cost_code_warnings)

        if dry_run:
            imported_count += 1
            continue

        try:
            client.insert(
                "budget_lines",
                {
                    "project_id": project_id,
                    "cost_code_id": cost_code_id,
                    "cost_type_id": cost_type_id,
                    "description": row.description or None,
                    "original_amount": row.original_budget,
                    "quantity": None,
                    "unit_of_measure": None,
                    "unit_cost": None,
                },
            )
            imported_count += 1
        except RuntimeError as exc:
            if "uq_budget_line" in str(exc) or '"code":"23505"' in str(exc):
                skipped_duplicates += 1
                warnings.append(
                    f'Row {index}: Skipped duplicate budget line for cost code "{cost_code_id}" and cost type "{row.cost_type}".'
                )
                continue
            raise

    total_budget, budget_warning = update_project_budget_summary(client, project_id)
    if budget_warning:
        warnings.append(
            "Project budget summary field was not updated cleanly. "
            f"Computed total is {total_budget:.2f}. Details: {budget_warning}"
        )

    return {
        "imported_count": imported_count,
        "skipped_duplicates": skipped_duplicates,
        "warnings": warnings,
        "total_budget": total_budget,
    }


def summarize_rows(rows: list[LegacyBudgetRow]) -> None:
    total_budget = sum(row.original_budget for row in rows)
    cost_type_counts: dict[str, int] = {}
    for row in rows:
        cost_type_counts[row.cost_type] = cost_type_counts.get(row.cost_type, 0) + 1

    print(f"Prepared {len(rows)} budget rows")
    print(f"Original budget total: {total_budget:.2f}")
    print(f"Cost type mix: {json.dumps(cost_type_counts, sort_keys=True)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, help="Legacy budget export (.xlsx)")
    parser.add_argument(
        "--project",
        help="Numeric project id or a unique project name/number substring",
    )
    parser.add_argument(
        "--import",
        dest="do_import",
        action="store_true",
        help="Import the converted rows into Supabase",
    )
    parser.add_argument(
        "--convert-only",
        action="store_true",
        help="Only write the normalized CSV; do not import",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Run all import checks without inserting budget_lines",
    )
    parser.add_argument(
        "--list-projects",
        action="store_true",
        help="List active projects and exit",
    )
    parser.add_argument(
        "--project-query",
        help="Optional filter used with --list-projects",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    client = get_supabase_client()

    if args.list_projects:
        list_projects(client, args.project_query)
        return 0

    if not args.input:
        raise SystemExit("--input is required unless --list-projects is used.")

    if not args.input.exists():
        raise SystemExit(f"Input file not found: {args.input}")

    rows = parse_legacy_workbook(args.input)
    summarize_rows(rows)
    output_csv = write_import_csv(rows, args.input)
    print(f"Wrote normalized CSV: {output_csv}")

    if args.convert_only or not args.do_import:
        return 0

    if not args.project:
        raise SystemExit("--project is required when --import is used.")

    project = resolve_project(client, args.project)
    project_id = int(project["id"])
    print(
        f'Import target: {project_id} - {project.get("project_number") or "-"} - {project.get("name") or "(unnamed)"}'
    )

    result = import_rows(client, project_id, rows, args.dry_run)
    mode = "Dry run complete" if args.dry_run else "Import complete"
    print(
        f"{mode}: {result['imported_count']} rows imported"
        + (
            f", {result.get('skipped_duplicates', 0)} duplicates skipped"
            if not args.dry_run
            else ""
        )
    )
    print(f"Project budget total after import: {result['total_budget']:.2f}")
    if result["warnings"]:
        print("Warnings:")
        for warning in result["warnings"]:
            print(f"  - {warning}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # pragma: no cover
        print(str(exc), file=sys.stderr)
        raise SystemExit(1) from exc
