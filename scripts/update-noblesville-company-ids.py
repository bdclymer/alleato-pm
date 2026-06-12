#!/usr/bin/env python3

import openpyxl
from pathlib import Path

# Company UUID mapping
COMPANY_MAPPING = {
    "C&S Heating & Cooling Inc.": "bb92bdf1-aa1f-46f7-b078-d851e5955302",
    "Casework Concepts LLC": "acd6b024-7ef9-4ce9-b9ea-575f6659989e",
    "Superior Contractors Inc.": "f48d62e8-8d91-4ba4-a975-3949dc442771",
    "Deem, LLC": "75e38d22-a809-42ac-99cf-c494d12b2b7e",
    "Justin Dorsey Plumbing": "c8c5a808-b214-4082-8f42-d0fe1f8bd97c",
    "KLEENIT GROUP INC.": "ff95148a-25b1-41b9-a26f-f841e4a73f44",
    "Inline Painting LLC": "238ec404-14c8-4954-bb3a-66f483f5e9e9",
    "Integrity Concrete Cutters": "0aaf7454-5db4-417c-b316-3f27c7b8437d",
    "Central Security & Communications": "e2877efa-d68e-4158-a422-124d1b1df73e",
    "Vulcan Fire Protection Services, Inc.": "2221ebff-14c4-442f-9df7-1b3a4c67976a",
    "Division 4 Masonry": "1def4648-e0c0-4cbc-ba75-ee3333c366ce",
    "Market Ready Interiors Inc": "d4b48e8a-db18-4f0b-8512-9efae4c81a8b",
    "Apex Glass LLC": "87f6a109-4889-4c30-a5ef-5a43eda76202",
    "Bul-Tec Roofing": "e8c6c563-6e80-4fcb-9e60-6a3bf80e7d51",
    "EURO Plastering": "a60e517c-badf-455e-9f20-acf7b2842493",
    "Steel Services, Inc.": "44f84efb-a5e7-4a86-a42f-b3d08300ab93",
    # "Central Indiana Hardware, CO, Inc": "[NOT FOUND - CREATE MANUALLY]",  # Not found - must be created first
}

# Vendor name lookup based on commitment numbers
VENDOR_BY_COMMITMENT = {
    "SC-5092-0001": "Integrity Concrete Cutters",
    "SC-5092-0002": "Superior Contractors Inc.",
    "SC-5092-0003": "Justin Dorsey Plumbing",
    "SC-5092-0004": "Market Ready Interiors Inc",
    "SC-5092-0005": "C&S Heating & Cooling Inc.",
    "SC-5092-0006": "Vulcan Fire Protection Services, Inc.",
    "SC-5092-0007": "Central Security & Communications",
    "SC-5092-0008": "Bul-Tec Roofing",
    "SC-5092-0009": "Deem, LLC",
    "SC-5092-0010": "Apex Glass LLC",
    "SC-5092-0011": "Central Indiana Hardware, CO, Inc",
    "SC-5092-0012": "Casework Concepts LLC",
    "SC-5092-0013": "Inline Painting LLC",
    "SC-5092-0014": "EURO Plastering",
    "SC-5092-0015": "Division 4 Masonry",
    "SC-5092-0016": "Steel Services, Inc.",
    "SC-5092-0017": "Central Indiana Hardware, CO, Inc",
    "SC-5092-0018": "KLEENIT GROUP INC.",
}

def update_company_ids():
    file_path = Path("./frontend/public/templates/noblesville-corrected-import.xlsx")

    if not file_path.exists():
        print(f"❌ File not found: {file_path}")
        return

    print(f"📖 Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path)

    # Find the sheet with commitment data
    sheet_name = "Commitments - Corrected" if "Commitments - Corrected" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    print(f"   Sheet: {sheet_name}")

    # Find the Company ID column (column E according to the template)
    # First, let's identify the header row
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value] = col_idx

    print(f"📋 Headers found: {list(headers.keys())}")

    company_id_col = headers.get("Company ID", 5)  # Default to column E
    commitment_number_col = headers.get("Number", 1)  # Default to column A

    print(f"🔄 Updating Company IDs in column {company_id_col}...")

    updated_count = 0
    missing_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=100), start=2):
        commitment_number = ws.cell(row_idx, commitment_number_col).value

        if not commitment_number:
            break  # End of data

        # Look up the vendor for this commitment
        vendor = VENDOR_BY_COMMITMENT.get(commitment_number)
        if not vendor:
            continue

        # Get the UUID for this vendor
        uuid = COMPANY_MAPPING.get(vendor)
        if not uuid:
            print(f"⚠️  {commitment_number}: Vendor '{vendor}' not found in mapping")
            missing_count += 1
            continue

        # Update the Company ID cell
        ws.cell(row_idx, company_id_col).value = uuid
        print(f"✅ {commitment_number}: {vendor}")
        updated_count += 1

    print(f"\n💾 Saving updated file...")
    wb.save(file_path)

    print(f"\n📊 Summary:")
    print(f"   Updated: {updated_count}")
    print(f"   Missing: {missing_count}")

    if missing_count > 0:
        print(f"\n⚠️  WARNING: {missing_count} commitments have missing company IDs")
        print("   You need to create these companies in Alleato before importing:")
        for commitment, vendor in VENDOR_BY_COMMITMENT.items():
            if vendor not in COMPANY_MAPPING:
                print(f"   - {vendor} (for {commitment})")

if __name__ == "__main__":
    update_company_ids()
